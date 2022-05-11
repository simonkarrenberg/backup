import pandas as pd
import datetime as dt
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from pyparsing import col
from sqlalchemy import column, false

## PARAMETERS

current_year = "2022"

## Load Datasets
core_data_monthly = pd.read_excel("C:/Users/LIJOW/OneDrive - LANXESS Deutschland GmbH/HR APP Daten/App Data/Core_Query_X_HTPHCHC3_CON_00_MONTHLY.xlsx",
sheet_name="Sheet1",
skiprows=1)

employee_mapping = pd.read_excel("C:/Users/LIJOW/OneDrive - LANXESS Deutschland GmbH/HR APP Daten/App Data/Employee_Mapping.xlsx",
sheet_name="Tabelle1")
employee_mapping = employee_mapping[["PID", "Manager for event reporting"]]

event_mapping = pd.read_excel("C:/Users/LIJOW/OneDrive - LANXESS Deutschland GmbH/HR APP Daten/App Data/Event_Mapping.xlsx",
sheet_name="my_HR_Events")

event_data_monthly = pd.read_excel("C:/Users/LIJOW/OneDrive - LANXESS Deutschland GmbH/HR APP Daten/App Data/X_HTPHCEV3_CON_001.xlsx",
sheet_name="Sheet1")
event_mapping = event_mapping[["Event Reason (myHR) - Code", "CT Event"]]
event_data_monthly = event_data_monthly.merge(event_mapping, how="left", \
                                              left_on="Event Reason (myHR)", \
                                              right_on="Event Reason (myHR) - Code")

hr_app_database = pd.read_excel("C:/Users/LIJOW/OneDrive - LANXESS Deutschland GmbH/HR APP Daten/Transformation/hr_app_database.xlsx",
sheet_name="Sheet1")
if "__PowerAppsId__" in hr_app_database.columns.values:
    hr_app_database = hr_app_database.drop(columns=["__PowerAppsId__", "id"])
else:
    hr_app_database = hr_app_database.drop(columns=["id"])




# Core Data Monthly
## Data cleansing
core_data_monthly = core_data_monthly.rename(columns={"Unnamed: 1": "Legal Entity text", "Unnamed: 6": "Cost Center text"})

## Data Enriching
### Add Manager
core_data_monthly = core_data_monthly.merge(employee_mapping, left_on="Person ID", right_on="PID", how="left")
core_data_monthly = core_data_monthly.drop(columns=["PID"])

### Create concatenated "joined" columns
core_data_monthly["LE joined"] = core_data_monthly["Legal Entity text"].astype(str) \
                                  + " (" \
                                  + core_data_monthly["Legal Entity"].astype(str) \
                                  + ")"

core_data_monthly["CC joined"] = core_data_monthly["Cost Center text"].astype(str) \
                                  + " - " \
                                  + core_data_monthly["Cost Center"].astype(str)

core_data_monthly["Name joined"] = core_data_monthly["Last Name"].astype(str) \
                                  + ", " \
                                  + core_data_monthly["First Name"].astype(str)


### Add Columns From Latest Reporting Month Core Table
core_data = hr_app_database
core_data["Order"] = core_data["Reporting Month"].astype(str).str[:2]
core_data["Order"] = core_data["Order"].str.replace('.','')
if core_data.empty == false:
    core_data = core_data[core_data["Order"] == max(core_data["Order"])]

cols_to_merge = []
for cols in hr_app_database.columns.values:
    if cols not in core_data_monthly.columns.values and cols not in ["CT Event", "Event Date", "HC_from_event", "FTE_from_Event", "timestamp", "Order"]:
        cols_to_merge.append(cols)

core_data = core_data[cols_to_merge]

core_data_monthly = core_data_monthly.merge(core_data, how="left", left_on="Person ID", right_on="Employee number")

### Add Events
#### Filter event data to reporting month
event_data_monthly["Event Year"] = event_data_monthly["Event Date"].str[6:10]
event_data_monthly = event_data_monthly[event_data_monthly["Event Year"] == current_year]
event_data_monthly["Reporting Month"] = event_data_monthly["Event Date"].str[3:5].astype(int)
event_data_monthly = event_data_monthly[event_data_monthly["Reporting Month"] == max(core_data_monthly["Reporting Month"].astype(str).str[:2].str.replace(".","").astype(int))]
event_data_monthly = event_data_monthly[["Person ID", "CT Event", "Event Date", "HC", "FTE"]]
event_data_monthly = event_data_monthly.rename(columns={"HC": "HC_from_event", "FTE": "FTE_from_Event"})
core_data_monthly = core_data_monthly.merge(event_data_monthly, how="left", on="Person ID")

### Add timestamp and prefill other columns
core_data_monthly["timestamp"] = core_data_monthly["Reporting Month"]
core_data_monthly["Position number"] = core_data_monthly["Position"]
core_data_monthly["Employee number"] = core_data_monthly["Person ID"]
core_data_monthly["Name"] = core_data_monthly["Last Name"]
core_data_monthly["LE number"] = core_data_monthly["Legal Entity"]
core_data_monthly["Cost center number"] = core_data_monthly["Cost Center"]
core_data_monthly["Date from"] = core_data_monthly["Event Date"]

### Set 'NA' according prefill logic by Hans Leo


hr_app_database = hr_app_database.drop(columns=["Order"])

core_data_final = hr_app_database.append(core_data_monthly)

core_data_final = core_data_final.drop_duplicates()
core_data_final['id'] = core_data_final.reset_index().index

## Save as excel file
time_string_cleaned = str(dt.datetime.now()).replace(":","").replace("-","").replace(".","")
backup_path = "C:/Users/LIJOW/OneDrive - LANXESS Deutschland GmbH/HR APP Daten/Transformation/backup/"
backup_name = "backup" + time_string_cleaned + ".xlsx"
core_data_final.to_excel(backup_path+backup_name, index=False)
core_data_final.to_excel('C:/Users/LIJOW/OneDrive - LANXESS Deutschland GmbH/HR APP Daten/Transformation/hr_app_database.xlsx', index=False)

## Convert to table within excel file
wb = load_workbook('C:/Users/LIJOW/OneDrive - LANXESS Deutschland GmbH/HR APP Daten/Transformation/hr_app_database.XLSX')
ws1 = wb["Sheet1"]

tab = Table(displayName="HR_APP_DATABASE", ref="A1:" + get_column_letter(ws1.max_column) + str(ws1.max_row))
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab.tableStyleInfo = style
ws1.add_table(tab)
wb.save("hr_app_database.XLSX")
